import oci
import pandas as pd
import gzip
import io
from scipy import stats
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# OCI configuration
config = oci.config.from_file()
object_storage = oci.object_storage.ObjectStorageClient(config)

namespace_name = 'ociateam'
bucket_name = 'cost_and_usage_reports'

object_list = object_storage.list_objects(namespace_name, bucket_name)
objects = object_list.data.objects

dataframes = []
missing_column_files = []

for obj in objects:
    object_name = obj.name
    
    # Process only files with a .gz extension
    if not object_name.endswith('.gz'):
        print(f"Skipping non-gz file: {object_name}")
        continue

    obj_data = object_storage.get_object(namespace_name, bucket_name, object_name)
    print(f"Processing Object {object_name}")
    
    # Read and process the .gz file
    with gzip.GzipFile(fileobj=io.BytesIO(obj_data.data.content)) as gz:
        try:
            df = pd.read_csv(gz, low_memory=False, usecols=['BillingPeriodEnd', 'BillingPeriodStart', 'EffectiveCost', 'Region', 'ServiceName'])

            # Optimize data types
            df['BillingPeriodStart'] = pd.to_datetime(df['BillingPeriodStart'], utc=True, errors='coerce')
            df['BillingPeriodEnd'] = pd.to_datetime(df['BillingPeriodEnd'], utc=True, errors='coerce')
            df['EffectiveCost'] = pd.to_numeric(df['EffectiveCost'], errors='coerce')
            df['Region'] = df['Region'].astype('category')
            df['ServiceName'] = df['ServiceName'].astype('category')
            
            # Check for required columns
            if df.columns.isin(['BillingPeriodEnd', 'BillingPeriodStart', 'EffectiveCost', 'Region', 'ServiceName']).all():
                dataframes.append(df)
            else:
                print(f"Missing required columns in file: {object_name}")
                missing_column_files.append(object_name)
                
        except Exception as e:
            print(f"Error processing file {object_name}: {e}")
            missing_column_files.append(object_name)

# Display files missing required columns
if missing_column_files:
    for file_name in missing_column_files:
        print(f"File with missing columns or error: {file_name}")

# Continue with further processing if valid data is found
if dataframes:
    final_df = pd.concat(dataframes, ignore_index=True)
    
    # Filter rows within the last 120 days
    current_time_utc = pd.Timestamp.now(tz='UTC')
    filtered_df = final_df.query("BillingPeriodStart >= @current_time_utc - pd.Timedelta(days=120)")

    # Proceed with further processing only if filtered_df is not empty
    if not filtered_df.empty:
        grouped_data = filtered_df.groupby(['ServiceName', 'Region'], as_index=False)['EffectiveCost'].sum()
        grouped_data['z_score'] = stats.zscore(grouped_data['EffectiveCost'])

        anomalies = grouped_data[np.abs(grouped_data['z_score']) > 3]

        # Processing anomalies further if they exist
        if not anomalies.empty:
            for index, row in anomalies.iterrows():
                service = row['ServiceName']
                region = row['Region']
                historical_data = final_df[(final_df['ServiceName'] == service) & (final_df['Region'] == region)]
                historical_data['month'] = historical_data['BillingPeriodStart'].dt.to_period('M')
                monthly_cost = historical_data.groupby('month')['EffectiveCost'].sum().reset_index()
                monthly_cost['pct_change'] = monthly_cost['EffectiveCost'].pct_change() * 100

                for i, month in enumerate(monthly_cost['month'].unique()):
                    month_cost = monthly_cost.loc[monthly_cost['month'] == month, 'EffectiveCost'].values[0]
                    pct_change = monthly_cost.loc[monthly_cost['month'] == month, 'pct_change'].values[0]
                    anomalies.loc[index, f"{month.strftime('%b')}_Cost"] = f"{month_cost:.2f}"
                    anomalies.loc[index, f"{month.strftime('%b')}_PctChange"] = pct_change

            # Save to Excel and apply conditional formatting
            anomaly_output_file = r'C:\Security\Blogs\Cost and Usage\Logs\usage_anomalies_with_history.xlsx'
            anomalies.to_excel(anomaly_output_file, index=False)

            # Apply conditional formatting in Excel
            wb = load_workbook(anomaly_output_file)
            ws = wb.active

            for col in ws.iter_cols(min_col=ws.max_column - len(monthly_cost['month'].unique()) + 1, max_col=ws.max_column):
                for cell in col:
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 0:
                            cell.font = Font(color="00FF00")  # Green for positive
                        elif cell.value < 0:
                            cell.font = Font(color="FF0000")  # Red for negative

            wb.save(anomaly_output_file)

            print(f"Anomalies with historical data and formatting saved to: {anomaly_output_file}")
        else:
            print("No anomalies found after filtering.")
    else:
        print("No valid data found within the last 120 days.")
