import oci
import pandas as pd
import gzip
import io
from sklearn.ensemble import IsolationForest
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font

# OCI configuration
config = oci.config.from_file()
object_storage = oci.object_storage.ObjectStorageClient(config)

namespace_name = 'ociateam'
bucket_name = 'cost_and_usage_reports'

object_list = object_storage.list_objects(namespace_name, bucket_name)
objects = object_list.data.objects

dataframes = []
missing_column_files = []

for obj in objects:
    object_name = obj.name
    obj_data = object_storage.get_object(namespace_name, bucket_name, object_name)
    print(f"Processing Object {object_name}")
    with gzip.GzipFile(fileobj=io.BytesIO(obj_data.data.content)) as gz:
        df = pd.read_csv(gz, low_memory=False)
        
        required_columns = ['BillingPeriodEnd', 'BillingPeriodStart', 'EffectiveCost', 'Region', 'ServiceName']
        if all(col in df.columns for col in required_columns):
            dataframes.append(df)
        else:
            missing_column_files.append(object_name)

if missing_column_files:
    for file_name in missing_column_files:
        print(file_name)

if dataframes:
    final_df = pd.concat(dataframes, ignore_index=True)
    final_df['BillingPeriodStart'] = pd.to_datetime(final_df['BillingPeriodStart'], utc=True, errors='coerce')
    final_df['BillingPeriodEnd'] = pd.to_datetime(final_df['BillingPeriodEnd'], utc=True, errors='coerce')

    current_time_utc = pd.Timestamp.now(tz='UTC')
    last_120_days = final_df[final_df['BillingPeriodStart'] >= current_time_utc - pd.Timedelta(days=120)]

    # Group by Service and Region for cost analysis
    grouped_data = last_120_days.groupby(['ServiceName', 'Region'])['EffectiveCost'].sum().reset_index()

    # Anomaly detection with Isolation Forest
    model = IsolationForest(contamination=0.05, random_state=42)
    grouped_data['anomaly'] = model.fit_predict(grouped_data[['EffectiveCost']])
    anomalies = grouped_data[grouped_data['anomaly'] == -1]

    # Adding monthly history and percentage change for each anomaly
    for index, row in anomalies.iterrows():
        service = row['ServiceName']
        region = row['Region']
        historical_data = final_df[(final_df['ServiceName'] == service) & (final_df['Region'] == region)]
        historical_data['month'] = historical_data['BillingPeriodStart'].dt.to_period('M')
        monthly_cost = historical_data.groupby('month')['EffectiveCost'].sum().reset_index()
        monthly_cost['pct_change'] = monthly_cost['EffectiveCost'].pct_change() * 100

        for i, month in enumerate(monthly_cost['month'].unique()):
            month_cost = monthly_cost.loc[monthly_cost['month'] == month, 'EffectiveCost'].values[0]
            pct_change = monthly_cost.loc[monthly_cost['month'] == month, 'pct_change'].values[0]
            anomalies.loc[index, f"{month.strftime('%b')}_Cost"] = f"{month_cost:.2f}"
            anomalies.loc[index, f"{month.strftime('%b')}_PctChange"] = pct_change

    # Save anomalies to Excel with conditional formatting
    anomaly_output_file = r'C:\Security\Blogs\Cost and Usage\Logs\isolation_usage_anomalies_with_history.xlsx'
    anomalies.to_excel(anomaly_output_file, index=False)

    # Apply conditional formatting in Excel
    wb = load_workbook(anomaly_output_file)
    ws = wb.active

    # Apply color formatting for percentage change columns
    for col in ws.iter_cols(min_col=ws.max_column - len(monthly_cost['month'].unique()), max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                if cell.value > 0:
                    cell.font = Font(color="00FF00")  # Green for positive
                elif cell.value < 0:
                    cell.font = Font(color="FF0000")  # Red for negative

    wb.save(anomaly_output_file)

    print(f"Anomalies with historical data and formatting saved to: {anomaly_output_file}")
else:
    print("No valid data found with required columns.")
